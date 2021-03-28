#user.py

from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_money = 3
c_lvl = 4

default_money = 10000

wb = load_workbook("userDB.xlsx")
ws = wb.active

def signup(_name, _id):
    ws.cell(row=2, column=c_name, value=_name)
    ws.cell(row=2, column=c_id, value =_id)
    ws.cell(row=2, column=c_money, value = default_money)
    ws.cell(row=2, column=c_lvl, value = 1)

    wb.save("userDB.xlsx")



   